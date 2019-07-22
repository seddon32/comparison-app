from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook('Mulsanne PC Rating Guide - April-May 2019 (CDL Only Spec).xlsx')
tci = Workbook('products.tcifile.xlsx')
ws = tci.active
ws1 = tci.create_sheet("Premium Breakdown", 0)

pb = wb.get_sheet_by_name('Premium Breakdown wef Apr 19')
Base = pb['A7':'A8']


print(wb.sheetnames)
print (tci.sheetnames)
print(pb)
print (Base)
